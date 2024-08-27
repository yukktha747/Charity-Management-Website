from django.views.generic import ListView, DetailView, View, FormView
from django.views.generic.edit import UpdateView, DeleteView, CreateView
from django.views.generic.detail import SingleObjectMixin
from django.urls import reverse_lazy, reverse
from django.shortcuts import get_object_or_404
from django.http import HttpResponse
from openpyxl import Workbook
from io import BytesIO
from .forms import CommentForm
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font
from django.utils import timezone
from django.contrib.auth.mixins import LoginRequiredMixin, UserPassesTestMixin
from django.contrib import messages
from django.shortcuts import redirect
from .models import (
    Case,
    Family_Member,
    Family_Expenses,
    Family_Income,
    Medical_Expenses,
    Notes,
    Regions,
    TypeHelp,
)
from django.urls import reverse
from .forms import CaseForm, Family_MemberForm


class AddRegionView(LoginRequiredMixin, UserPassesTestMixin, CreateView):
    model = Regions
    fields = ["region", "city"]
    template_name = "add_region.html"
    success_url = reverse_lazy("case_list")

    def test_func(self):
        return self.request.user.is_superuser

    def handle_no_permission(self):
        messages.error(self.request, "You do not have permission to add a region.")
        return redirect("case_list")


class AddHelpView(LoginRequiredMixin, UserPassesTestMixin, CreateView):
    model = TypeHelp
    fields = ["typeHelp"]
    template_name = "add_help.html"
    success_url = reverse_lazy("case_list")

    def test_func(self):
        return self.request.user.is_superuser

    def handle_no_permission(self):
        messages.error(self.request, "You do not have permission to add an aid type.")
        return redirect("case_list")


class CommentGet(LoginRequiredMixin, DetailView):
    model = Case
    template_name = "case_detail.html"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["form"] = CommentForm()
        return context

    def form_valid(self, form):
        comment = form.save(commit=False)
        comment.author = self.request.user
        comment.case = self.object
        comment.save()
        return super().form_valid(form)


class CommentPost(SingleObjectMixin, FormView):  # new
    model = Case
    form_class = CommentForm
    template_name = "case_detail.html"

    def post(self, request, *args, **kwargs):
        self.object = self.get_object()
        return super().post(request, *args, **kwargs)

    def form_valid(self, form):
        comment = form.save(commit=False)
        comment.author = self.request.user
        comment.case = self.object
        comment.save()
        return super().form_valid(form)

    def get_success_url(self):
        article = self.get_object()
        return reverse("case_detail", kwargs={"pk": article.pk})


class CaseSearchView(ListView):
    model = Case
    template_name = "case_search.html"
    context_object_name = "case_list"
    paginate_by = 10

    def get_queryset(self):
        query = self.request.GET.get("query")
        if query:
            return Case.objects.filter(name__icontains=query)
        return Case.objects.none()


class CaseListView(LoginRequiredMixin, ListView):
    model = Case
    template_name = "case_list.html"


class CaseDetailView(LoginRequiredMixin, DetailView):
    model = Case
    template_name = "case_detail.html"
    context_object_name = "case"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        case = self.get_object()
        family_members = Family_Member.objects.filter(case=case)
        family_income = Family_Income.objects.filter(case=case)
        family_expenses = Family_Expenses.objects.filter(case=case)
        medical_expenses = Medical_Expenses.objects.filter(case=case)
        notes = Notes.objects.filter(case=case)
        context["family_members"] = family_members
        context["family_income"] = family_income
        context["family_expenses"] = family_expenses
        context["medical_expenses"] = medical_expenses
        context["notes"] = notes
        context["form"] = CommentForm()
        return context

    def get(self, request, *args, **kwargs):
        view = CommentGet.as_view()
        return view(request, *args, **kwargs)

    def post(self, request, *args, **kwargs):
        view = CommentPost.as_view()
        return view(request, *args, **kwargs)


class CaseUpdateView(LoginRequiredMixin, UpdateView):
    model = Case
    fields = (
        "name",
        "gender",
        "job",
        "region",
        "typeHelp",
        "marriageStatus",
        "birthDate",
        "nationalID",
        "nationalIDExpiration",
        "qualification",
        "phoneNumber",
        "gaurdianName",
        "gaurdianRelation",
        "gaudrianNumber",
        "housing",
        "caseDescribtion",
    )
    template_name = "case_edit.html"


class CaseDeleteView(LoginRequiredMixin, DeleteView):
    model = Case
    template_name = "case_delete.html"
    success_url = reverse_lazy("case_list")


class CaseCreateView(LoginRequiredMixin, CreateView):
    model = Case
    template_name = "case_new.html"
    form_class = CaseForm

    def form_valid(self, form):
        form.instance.author = self.request.user
        return super().form_valid(form)


class FamilyMemberCreateView(LoginRequiredMixin, CreateView):
    model = Family_Member
    template_name = "family_members/member_new.html"
    form_class = Family_MemberForm

    def get_success_url(self):
        return reverse(
            "case_detail", kwargs={"pk": self.kwargs["pk"]}
        )  # Provide the 'pk' argument

    def form_valid(self, form):
        case = Case.objects.get(pk=self.kwargs["pk"])
        form.instance.case = case
        return super().form_valid(form)


class FamilyMemberUpdateView(LoginRequiredMixin, UpdateView):
    model = Family_Member
    fields = (
        "name",
        "gender",
        "age",
        "qualification",
        "occupation",
        "notes",
    )
    template_name = "family_members/member_edit.html"

    def get_success_url(self):
        return reverse("case_detail", kwargs={"pk": self.kwargs["case_pk"]})

    def form_valid(self, form):
        case = Case.objects.get(pk=self.kwargs["case_pk"])
        form.instance.case = case
        return super().form_valid(form)


class FamilyMemberDeleteView(LoginRequiredMixin, DeleteView):
    model = Family_Member
    template_name = "family_members/member_delete.html"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        case = get_object_or_404(Case, pk=self.kwargs["case_pk"])
        context["case"] = case
        return context

    def get_success_url(self):
        return reverse("case_detail", kwargs={"pk": self.kwargs["case_pk"]})


class FamilyIncomeCreateView(LoginRequiredMixin, CreateView):
    model = Family_Income
    template_name = "family_income/income_new.html"
    fields = (
        "source_name",
        "amount",
    )

    def get_success_url(self):
        return reverse("case_detail", kwargs={"pk": self.kwargs["pk"]})

    def form_valid(self, form):
        case = Case.objects.get(pk=self.kwargs["pk"])
        form.instance.case = case
        return super().form_valid(form)


class FamilyIncomeUpdateView(LoginRequiredMixin, UpdateView):
    model = Family_Income
    fields = (
        "source_name",
        "amount",
    )
    template_name = "family_income/income_edit.html"

    def get_success_url(self):
        return reverse("case_detail", kwargs={"pk": self.kwargs["case_pk"]})

    def form_valid(self, form):
        case = Case.objects.get(pk=self.kwargs["case_pk"])
        form.instance.case = case
        return super().form_valid(form)


class FamilyIncomeDeleteView(LoginRequiredMixin, DeleteView):
    model = Family_Income
    template_name = "family_income/income_delete.html"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        income = get_object_or_404(Family_Income, pk=self.kwargs["pk"])
        case = income.case
        context["income"] = income
        context["case"] = case
        return context

    def get_success_url(self):
        return reverse("case_detail", kwargs={"pk": self.kwargs["case_pk"]})


class FamilyExpensesCreateView(LoginRequiredMixin, CreateView):
    model = Family_Expenses
    template_name = "family_expenses/expense_new.html"
    fields = (
        "source_name",
        "amount",
    )

    def get_success_url(self):
        return reverse("case_detail", kwargs={"pk": self.kwargs["pk"]})

    def form_valid(self, form):
        case = Case.objects.get(pk=self.kwargs["pk"])
        form.instance.case = case
        return super().form_valid(form)


class FamilyExpensesUpdateView(LoginRequiredMixin, UpdateView):
    model = Family_Expenses
    fields = (
        "source_name",
        "amount",
    )
    template_name = "family_expenses/expense_edit.html"

    def get_success_url(self):
        return reverse("case_detail", kwargs={"pk": self.kwargs["case_pk"]})

    def form_valid(self, form):
        case = Case.objects.get(pk=self.kwargs["case_pk"])
        form.instance.case = case
        return super().form_valid(form)


class FamilyExpensesDeleteView(LoginRequiredMixin, DeleteView):
    model = Family_Expenses
    template_name = "family_expenses/expense_delete.html"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        expense = get_object_or_404(Family_Expenses, pk=self.kwargs["pk"])
        case = expense.case
        context["expense"] = expense
        context["case"] = case
        return context

    def get_success_url(self):
        return reverse("case_detail", kwargs={"pk": self.kwargs["case_pk"]})


class MedicalExpensesCreateView(LoginRequiredMixin, CreateView):
    model = Medical_Expenses
    template_name = "family_medical/medical_new.html"
    fields = (
        "fullName",
        "diseaseType",
        "medicine",
        "insuranceID",
    )

    def get_success_url(self):
        return reverse("case_detail", kwargs={"pk": self.kwargs["pk"]})

    def form_valid(self, form):
        case = Case.objects.get(pk=self.kwargs["pk"])
        form.instance.case = case
        return super().form_valid(form)


class MedicalExpensesUpdateView(LoginRequiredMixin, UpdateView):
    model = Medical_Expenses
    fields = (
        "fullName",
        "diseaseType",
        "medicine",
        "insuranceID",
    )
    template_name = "family_medical/medical_edit.html"

    def get_success_url(self):
        return reverse("case_detail", kwargs={"pk": self.kwargs["case_pk"]})

    def form_valid(self, form):
        case = Case.objects.get(pk=self.kwargs["case_pk"])
        form.instance.case = case
        return super().form_valid(form)


class MedicalExpensesDeleteView(LoginRequiredMixin, DeleteView):
    model = Medical_Expenses
    template_name = "family_medical/medical_delete.html"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        medical = get_object_or_404(Medical_Expenses, pk=self.kwargs["pk"])
        case = medical.case
        context["medical"] = medical
        context["case"] = case
        return context

    def get_success_url(self):
        return reverse("case_detail", kwargs={"pk": self.kwargs["case_pk"]})


class NoteCreateView(LoginRequiredMixin, CreateView):
    model = Notes
    template_name = "notes/note_new.html"
    fields = (
        "noteHeader",
        "humanNeeds",
        "interviewDescription",
        "interviewResult",
        "researcherOpinion",
        "supervisorOpinion",
        "overallRating",
    )

    def get_success_url(self):
        return reverse("case_detail", kwargs={"pk": self.kwargs["pk"]})

    def form_valid(self, form):
        case = Case.objects.get(pk=self.kwargs["pk"])
        form.instance.case = case
        return super().form_valid(form)


class NoteUpdateView(LoginRequiredMixin, UpdateView):
    model = Notes
    fields = (
        "noteHeader",
        "humanNeeds",
        "interviewDescription",
        "interviewResult",
        "researcherOpinion",
        "supervisorOpinion",
        "overallRating",
    )
    template_name = "notes/note_edit.html"

    def get_success_url(self):
        return reverse("case_detail", kwargs={"pk": self.kwargs["case_pk"]})

    def form_valid(self, form):
        case = Case.objects.get(pk=self.kwargs["case_pk"])
        form.instance.case = case
        return super().form_valid(form)


class NoteDeleteView(LoginRequiredMixin, DeleteView):
    model = Notes
    template_name = "notes/note_delete.html"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        note = get_object_or_404(Notes, pk=self.kwargs["pk"])
        case = note.case
        context["note"] = note
        context["case"] = case
        return context

    def get_success_url(self):
        return reverse("case_detail", kwargs={"pk": self.kwargs["case_pk"]})


class DownloadExcelView(LoginRequiredMixin, View):
    def get(self, request, case_id):
        # Fetch the case from the database
        case = Case.objects.get(pk=case_id)
        if not request.user.is_superuser:
            messages.error(
                request, "You do not have permission to download the dataset."
            )
            return redirect(reverse("case_detail", kwargs={"pk": case_id}))
        try:
            case = Case.objects.get(pk=case_id)
        except Case.DoesNotExist:
            return HttpResponse("Case not found", status=404)

        # Create a new workbook
        workbook = Workbook()

        # Create sheets for each table
        case_sheet = workbook.active
        case_sheet.title = "Case"
        family_members_sheet = workbook.create_sheet(title="Family Members")
        family_expenses_sheet = workbook.create_sheet(title="Family Expenses")
        family_income_sheet = workbook.create_sheet(title="Family Income")
        medical_expenses_sheet = workbook.create_sheet(title="Medical Expenses")
        notes_sheet = workbook.create_sheet(title="Notes")

        # Add headers to the Case sheet
        case_sheet.append(
            [
                "ID",
                "Name",
                "Gender",
                "Marriage Status",
                "Birth Date",
                "National ID",
                "Type Help",
                "Region",
            ]
        )
        case_sheet["A1"].font = Font(bold=True)
        for col in range(2, 9):
            column_letter = get_column_letter(col)
            case_sheet[column_letter + "1"].font = Font(bold=True)

        # Add data to the Case sheet
        case_sheet.append(
            [
                case.pk,
                case.name,
                case.get_gender_display(),
                case.get_marriageStatus_display(),
                case.birthDate,
                case.nationalID,
                case.typeHelp.typeHelp if case.typeHelp else "",
                case.region.region if case.region else "",
            ]
        )

        # Add headers to the Family Members sheet
        family_members_sheet.append(
            ["Case ID", "Name", "Gender", "Age", "Qualification", "Occupation", "Notes"]
        )
        family_members_sheet["A1"].font = Font(bold=True)
        for col in range(2, 8):
            column_letter = get_column_letter(col)
            family_members_sheet[column_letter + "1"].font = Font(bold=True)

        # Add data to the Family Members sheet
        for member in case.family_members.all():
            family_members_sheet.append(
                [
                    case_id,
                    member.name,
                    member.gender,
                    member.age,
                    member.qualification,
                    member.occupation,
                    member.notes,
                ]
            )

        # Add headers to the Family Expenses sheet
        family_expenses_sheet.append(["Case ID", "Statement", "Amount", "Notes"])
        family_expenses_sheet["A1"].font = Font(bold=True)
        for col in range(2, 5):
            column_letter = get_column_letter(col)
            family_expenses_sheet[column_letter + "1"].font = Font(bold=True)

        # Add data to the Family Expenses sheet
        for expense in case.family_expenses.all():
            family_expenses_sheet.append(
                [case_id, expense.statement, expense.amount, expense.notes]
            )

        # Add headers to the Family Income sheet
        family_income_sheet.append(["Case ID", "Source", "Amount"])
        family_income_sheet["A1"].font = Font(bold=True)
        for col in range(2, 4):
            column_letter = get_column_letter(col)
            family_income_sheet[column_letter + "1"].font = Font(bold=True)

        # Add data to the Family Income sheet
        for income in case.family_income.all():
            family_income_sheet.append([case_id, income.source_name, income.amount])
        # Add headers to the Medical Expenses sheet
        medical_expenses_sheet.append(
            ["Case ID", "Full Name", "Disease Type", "Medicine", "Insurance ID"]
        )
        medical_expenses_sheet["A1"].font = Font(bold=True)
        for col in range(2, 6):
            column_letter = get_column_letter(col)
            medical_expenses_sheet[column_letter + "1"].font = Font(bold=True)

        # Add data to the Medical Expenses sheet
        for expense in case.medical_expenses.all():
            medical_expenses_sheet.append(
                [
                    case_id,
                    expense.fullName,
                    expense.diseaseType,
                    expense.medicine,
                    expense.insuranceID,
                ]
            )

        # Add headers to the Notes sheet
        notes_sheet.append(
            [
                "Case ID",
                "Note Header",
                "Human Needs",
                "Other Help",
                "Interview Description",
                "Interview Result",
                "Researcher Opinion",
                "Supervisor Opinion",
                "Overall Rating",
            ]
        )
        notes_sheet["A1"].font = Font(bold=True)
        for col in range(2, 10):
            column_letter = get_column_letter(col)
            notes_sheet[column_letter + "1"].font = Font(bold=True)

        # Add data to the Notes sheet
        for note in case.notes.all():
            notes_sheet.append(
                [
                    case_id,
                    note.noteHeader,
                    note.humanNeeds,
                    note.otherHelp,
                    note.interviewDescription,
                    note.interviewResult,
                    note.researcherOpinion,
                    note.supervisorOpinion,
                    note.overallRating,
                ]
            )

        # Create a response object with the Excel file content type
        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

        # Set the file name
        response["Content-Disposition"] = 'attachment; filename="case_data.xlsx"'

        # Save the workbook to the response
        workbook.save(response)

        return response


class DownloadAllView(LoginRequiredMixin, View):
    def get(self, request):
        if not request.user.is_superuser:
            messages.error(
                request, "You do not have permission to download the dataset."
            )
            return redirect(reverse("case_list"))

        # Fetch all cases from the database
        cases = Case.objects.all()

        # Create a new workbook
        workbook = Workbook()

        # Create sheets for each table
        case_sheet = workbook.active
        case_sheet.title = "Case"
        family_members_sheet = workbook.create_sheet(title="Family Members")
        family_expenses_sheet = workbook.create_sheet(title="Family Expenses")
        family_income_sheet = workbook.create_sheet(title="Family Income")
        medical_expenses_sheet = workbook.create_sheet(title="Medical Expenses")
        notes_sheet = workbook.create_sheet(title="Notes")

        # Add headers to the Case sheet
        case_sheet.append(
            [
                "ID",
                "Name",
                "Gender",
                "Marriage Status",
                "Birth Date",
                "National ID",
                "Type Help",
                "Region",
            ]
        )
        case_sheet["A1"].font = Font(bold=True)
        for col in range(2, 9):
            column_letter = get_column_letter(col)
            case_sheet[column_letter + "1"].font = Font(bold=True)

        # Add headers to the Family Members sheet
        family_members_sheet.append(
            ["Case ID", "Name", "Gender", "Age", "Qualification", "Occupation", "Notes"]
        )
        family_members_sheet["A1"].font = Font(bold=True)
        for col in range(2, 8):
            column_letter = get_column_letter(col)
            family_members_sheet[column_letter + "1"].font = Font(bold=True)

        # Add headers to the Family Expenses sheet
        family_expenses_sheet.append(["Case ID", "Statement", "Amount", "Notes"])
        family_expenses_sheet["A1"].font = Font(bold=True)
        for col in range(2, 5):
            column_letter = get_column_letter(col)
            family_expenses_sheet[column_letter + "1"].font = Font(bold=True)

        # Add headers to the Family Income sheet
        family_income_sheet.append(["Case ID", "Source", "Amount"])
        family_income_sheet["A1"].font = Font(bold=True)
        for col in range(2, 4):
            column_letter = get_column_letter(col)
            family_income_sheet[column_letter + "1"].font = Font(bold=True)

        # Add headers to the Medical Expenses sheet
        medical_expenses_sheet.append(
            ["Case ID", "Full Name", "Disease Type", "Medicine", "Insurance ID"]
        )
        medical_expenses_sheet["A1"].font = Font(bold=True)
        for col in range(2, 6):
            column_letter = get_column_letter(col)
            medical_expenses_sheet[column_letter + "1"].font = Font(bold=True)

        # Add headers to the Notes sheet
        notes_sheet.append(
            [
                "Case ID",
                "Note Header",
                "Human Needs",
                "Other Help",
                "Interview Description",
                "Interview Result",
                "Researcher Opinion",
                "Supervisor Opinion",
                "Overall Rating",
            ]
        )
        notes_sheet["A1"].font = Font(bold=True)
        for col in range(2, 10):
            column_letter = get_column_letter(col)
            notes_sheet[column_letter + "1"].font = Font(bold=True)

            # Create a response object with the Excel file content type
            # Fetch all cases from the database
        cases = Case.objects.all()

        # Iterate over all cases
        for case in cases:
            # Add data to the Case sheet
            case_sheet.append(
                [
                    case.pk,
                    case.name,
                    case.get_gender_display(),
                    case.get_marriageStatus_display(),
                    case.birthDate,
                    case.nationalID,
                    case.typeHelp.typeHelp if case.typeHelp else "",
                    case.region.region if case.region else "",
                ]
            )

            # Add data to the Family Members sheet
            for member in case.family_members.all():
                family_members_sheet.append(
                    [
                        case.pk,
                        member.name,
                        member.gender,
                        member.age,
                        member.qualification,
                        member.occupation,
                        member.notes,
                    ]
                )

            # Add data to the Family Expenses sheet
            for expense in case.family_expenses.all():
                family_expenses_sheet.append(
                    [case.pk, expense.statement, expense.amount, expense.notes]
                )

            # Add data to the Family Income sheet
            for income in case.family_income.all():
                family_income_sheet.append([case.pk, income.source_name, income.amount])

            # Add data to the Medical Expenses sheet
            for expense in case.medical_expenses.all():
                medical_expenses_sheet.append(
                    [
                        case.pk,
                        expense.fullName,
                        expense.diseaseType,
                        expense.medicine,
                        expense.insuranceID,
                    ]
                )

            # Add data to the Notes sheet
            for note in case.notes.all():
                notes_sheet.append(
                    [
                        case.pk,
                        note.noteHeader,
                        note.humanNeeds,
                        note.otherHelp,
                        note.interviewDescription,
                        note.interviewResult,
                        note.researcherOpinion,
                        note.supervisorOpinion,
                        note.overallRating,
                    ]
                )

        # Create a response object with the Excel file content type
        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

        # Set the file name
        response["Content-Disposition"] = 'attachment; filename="case_data.xlsx"'

        # Save the workbook to the response
        workbook.save(response)

        return response
